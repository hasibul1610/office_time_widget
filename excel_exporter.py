"""
Excel Exporter module for Office Time Widget.
Generates beautifully formatted .xlsx workbooks with executive summaries,
KPI metric cards, weekly breakdowns, and daily time logs using openpyxl.
"""

from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, Optional

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from config import config
from db import DatabaseManager, db as global_db


class ExcelReportExporter:
    """Generates corporate-styled Excel monthly timesheet reports."""

    def __init__(self, database_manager: Optional[DatabaseManager] = None):
        self.db = database_manager or global_db

    @staticmethod
    def _format_seconds_to_hm(seconds: float) -> str:
        """Converts seconds into 'Xh Ym' or '0h 00m'."""
        total_mins = int(round(seconds / 60.0))
        hours = total_mins // 60
        mins = total_mins % 60
        return f"{hours}h {mins:02d}m"

    def export_monthly_report(
        self, year: int, month: int, output_path: Optional[Path] = None
    ) -> Path:
        """
        Builds and saves a full monthly report workbook.
        Returns the saved file Path.
        """
        month_data = self.db.get_month_summary(year, month)
        month_name = month_data["month_name"]
        daily_target_hours = float(config.get("daily_target_hours", 8.0))
        daily_min_hours = float(config.get("daily_min_hours", 4.0))
        weekly_target_hours = float(config.get("weekly_target_hours", 36.0))

        if output_path is None:
            from config import get_data_dir

            reports_dir = get_data_dir() / "reports"
            reports_dir.mkdir(parents=True, exist_ok=True)
            output_path = reports_dir / f"Time_Report_{year}_{month:02d}_{month_name}.xlsx"

        wb = openpyxl.Workbook()
        # Remove default sheet
        default_sheet = wb.active

        # Palette colors
        C_HEADER_BG = "1E293B"       # Slate 800
        C_HEADER_FG = "FFFFFF"       # White
        C_SUBHEADER_BG = "334155"    # Slate 700
        C_CARD_BG = "F1F5F9"         # Slate 100
        C_ZEBRA = "F8FAFC"           # Slate 50
        C_GREEN_BG = "DCFCE7"        # Emerald 100
        C_GREEN_FG = "166534"        # Emerald 800
        C_YELLOW_BG = "FEF9C3"       # Amber 100
        C_YELLOW_FG = "854D0E"       # Amber 800
        C_RED_BG = "FEE2E2"          # Rose 100
        C_RED_FG = "991B1B"          # Rose 800
        C_ACCENT_BLUE = "2563EB"     # Blue 600

        thin_border_side = Side(style="thin", color="CBD5E1")
        border_all = Border(
            left=thin_border_side,
            right=thin_border_side,
            top=thin_border_side,
            bottom=thin_border_side,
        )
        border_top_thick = Border(
            top=Side(style="medium", color="1E293B"),
            bottom=Side(style="double", color="1E293B"),
            left=thin_border_side,
            right=thin_border_side,
        )

        font_title = Font(name="Segoe UI", size=16, bold=True, color="1E293B")
        font_subtitle = Font(name="Segoe UI", size=10, italic=True, color="64748B")
        font_header = Font(name="Segoe UI", size=11, bold=True, color=C_HEADER_FG)
        font_bold = Font(name="Segoe UI", size=10, bold=True, color="1E293B")
        font_regular = Font(name="Segoe UI", size=10, color="1E293B")
        font_kpi_num = Font(name="Segoe UI", size=14, bold=True, color=C_ACCENT_BLUE)
        font_kpi_label = Font(name="Segoe UI", size=9, bold=True, color="64748B")

        # ----------------------------------------------------
        # 1. SHEET 1: MONTHLY SUMMARY
        # ----------------------------------------------------
        ws_sum = wb.create_sheet(title="Monthly Summary")
        ws_sum.views.sheetView[0].showGridLines = True

        # Title Block
        ws_sum.merge_cells("B2:H2")
        ws_sum["B2"] = "OFFICE TIME TRACKER - MONTHLY REPORT"
        ws_sum["B2"].font = font_title
        ws_sum["B2"].alignment = Alignment(vertical="center")

        ws_sum["B3"] = f"Period: {month_name} {year}   |   Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws_sum["B3"].font = font_subtitle

        # Configuration Parameters Box
        ws_sum["B5"] = "Configured Targets:"
        ws_sum["B5"].font = font_bold
        ws_sum["B6"] = f"• Daily Target: {daily_target_hours:.1f} hrs"
        ws_sum["B6"].font = font_regular
        ws_sum["D6"] = f"• Daily Minimum: {daily_min_hours:.1f} hrs"
        ws_sum["D6"].font = font_regular
        ws_sum["F6"] = f"• Weekly Target: {weekly_target_hours:.1f} hrs"
        ws_sum["F6"].font = font_regular

        # KPI Metric Cards (Rows 8-10)
        kpi_metrics = [
            ("TOTAL WORKED", f"{month_data['total_month_work_hours']:.1f} hrs", "B8:C9"),
            ("EXPECTED TARGET", f"{month_data['expected_target_hours']:.1f} hrs", "D8:E9"),
            (
                "OVERTIME / VARIANCE",
                f"{'+' if month_data['balance_hours'] >= 0 else ''}{month_data['balance_hours']:.1f} hrs",
                "F8:G9",
            ),
            ("ACTIVE DAYS", f"{month_data['active_days_count']} / {month_data['weekday_count']} days", "H8:I9"),
        ]

        for label, val_text, cell_range in kpi_metrics:
            start_cell = cell_range.split(":")[0]
            ws_sum.merge_cells(cell_range)
            cell = ws_sum[start_cell]
            cell.value = f"{label}\n{val_text}"
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.font = font_kpi_num

            # Color highlight variance
            if "OVERTIME" in label:
                if month_data["balance_hours"] >= 0:
                    fill_col = C_GREEN_BG
                    font_col = C_GREEN_FG
                else:
                    fill_col = C_RED_BG
                    font_col = C_RED_FG
                cell.font = Font(name="Segoe UI", size=13, bold=True, color=font_col)
                card_fill = PatternFill(start_color=fill_col, end_color=fill_col, fill_type="solid")
            else:
                card_fill = PatternFill(start_color=C_CARD_BG, end_color=C_CARD_BG, fill_type="solid")

            # Apply fill & border to card range
            r1, c1_letter = int(start_cell[1:]), start_cell[0]
            end_cell = cell_range.split(":")[1]
            r2, c2_letter = int(end_cell[1:]), end_cell[0]
            c1 = openpyxl.utils.column_index_from_string(c1_letter)
            c2 = openpyxl.utils.column_index_from_string(c2_letter)

            for r in range(r1, r2 + 1):
                for c in range(c1, c2 + 1):
                    cl = ws_sum.cell(row=r, column=c)
                    cl.fill = card_fill
                    cl.border = border_all

        # Weekly Breakdown Table (Starting Row 12)
        ws_sum["B12"] = "Weekly Breakdown"
        ws_sum["B12"].font = Font(name="Segoe UI", size=12, bold=True, color="1E293B")

        headers_week = ["Week #", "Date Range", "Worked (hrs)", "Target (hrs)", "Variance (hrs)", "Weekly Status"]
        for col_idx, h_text in enumerate(headers_week, start=2):
            cell = ws_sum.cell(row=13, column=col_idx, value=h_text)
            cell.fill = PatternFill(start_color=C_HEADER_BG, end_color=C_HEADER_BG, fill_type="solid")
            cell.font = font_header
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border_all
        ws_sum.row_dimensions[13].height = 24

        curr_r = 14
        for w in month_data["weeks"]:
            w_num = f"Week {w['week_num']}"
            w_range = w["range_label"]
            w_hours = round(w["work_hours"], 2)
            w_target = round(w["target_hours"], 2)
            w_var = round(w_hours - w_target, 2)
            w_status = "Target Met" if w_hours >= w_target else "Under Target"

            ws_sum.cell(row=curr_r, column=2, value=w_num).alignment = Alignment(horizontal="center")
            ws_sum.cell(row=curr_r, column=3, value=w_range).alignment = Alignment(horizontal="center")
            ws_sum.cell(row=curr_r, column=4, value=w_hours).number_format = "0.00"
            ws_sum.cell(row=curr_r, column=5, value=w_target).number_format = "0.00"
            
            var_cell = ws_sum.cell(row=curr_r, column=6, value=w_var)
            var_cell.number_format = "+0.00;-0.00;0.00"
            
            st_cell = ws_sum.cell(row=curr_r, column=7, value=w_status)
            st_cell.alignment = Alignment(horizontal="center")

            row_fill = PatternFill(
                start_color=C_ZEBRA if curr_r % 2 == 0 else "FFFFFF",
                end_color=C_ZEBRA if curr_r % 2 == 0 else "FFFFFF",
                fill_type="solid",
            )
            for c in range(2, 8):
                cl = ws_sum.cell(row=curr_r, column=c)
                cl.font = font_regular
                cl.border = border_all
                cl.fill = row_fill

            if w_status == "Target Met":
                st_cell.fill = PatternFill(start_color=C_GREEN_BG, end_color=C_GREEN_BG, fill_type="solid")
                st_cell.font = Font(name="Segoe UI", size=10, bold=True, color=C_GREEN_FG)
            else:
                st_cell.fill = PatternFill(start_color=C_YELLOW_BG, end_color=C_YELLOW_BG, fill_type="solid")
                st_cell.font = Font(name="Segoe UI", size=10, bold=True, color=C_YELLOW_FG)

            curr_r += 1

        # ----------------------------------------------------
        # 2. SHEET 2: DAILY TIME LOGS
        # ----------------------------------------------------
        ws_log = wb.create_sheet(title="Daily Time Logs")
        ws_log.views.sheetView[0].showGridLines = True

        ws_log.merge_cells("A1:K1")
        ws_log["A1"] = f"DETAILED DAILY LOGS - {month_name.upper()} {year}"
        ws_log["A1"].font = font_title
        ws_log["A1"].alignment = Alignment(vertical="center")
        ws_log.row_dimensions[1].height = 28

        headers_daily = [
            "Date",
            "Day",
            "First In",
            "Last Out",
            "Sessions",
            "Break (min)",
            "Work (hh:mm)",
            "Work (hrs)",
            "Target (hrs)",
            "Variance (hrs)",
            "Status",
            "Notes",
        ]

        for col_idx, h_text in enumerate(headers_daily, start=1):
            cell = ws_log.cell(row=3, column=col_idx, value=h_text)
            cell.fill = PatternFill(start_color=C_HEADER_BG, end_color=C_HEADER_BG, fill_type="solid")
            cell.font = font_header
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border_all
        ws_log.row_dimensions[3].height = 24

        log_r = 4
        daily_items = sorted(month_data["daily_data"].items(), key=lambda x: x[0])

        for date_str, d_info in daily_items:
            day_num = d_info["day"]
            day_name = d_info["day_name"]
            work_days = config.get("work_days", [5, 6, 0, 1, 2])
            is_weekend = d_info["weekday"] not in work_days
            sessions = d_info["sessions"]
            work_sec = d_info["work_sec"]
            break_sec = d_info["break_sec"]
            work_hrs = round(work_sec / 3600.0, 2)
            break_mins = int(round(break_sec / 60.0))

            first_in = "--:--"
            last_out = "--:--"
            notes_list = []

            if sessions:
                # Find earliest clock in
                valid_ins = [s["clock_in"] for s in sessions if s.get("clock_in")]
                if valid_ins:
                    first_in = datetime.fromisoformat(min(valid_ins)).strftime("%H:%M")

                valid_outs = [s["clock_out"] for s in sessions if s.get("clock_out")]
                if valid_outs:
                    last_out = datetime.fromisoformat(max(valid_outs)).strftime("%H:%M")
                elif any(s.get("status") in ("working", "break") for s in sessions):
                    last_out = "Active"

                for s in sessions:
                    if s.get("notes") and s["notes"].strip():
                        notes_list.append(s["notes"].strip())

            day_target = 0.0 if is_weekend else daily_target_hours
            variance = round(work_hrs - day_target, 2)

            if is_weekend and work_hrs == 0.0:
                status_text = "Weekend"
                status_cat = "weekend"
            elif work_hrs >= daily_target_hours:
                status_text = "Target Met (8h+)"
                status_cat = "green"
            elif work_hrs >= daily_min_hours:
                status_text = "Min Met (4h+)"
                status_cat = "yellow"
            elif work_hrs > 0:
                status_text = "Under Min (<4h)"
                status_cat = "red"
            else:
                status_text = "Off / Absent"
                status_cat = "red" if not is_weekend else "weekend"

            ws_log.cell(row=log_r, column=1, value=date_str).alignment = Alignment(horizontal="center")
            ws_log.cell(row=log_r, column=2, value=day_name).alignment = Alignment(horizontal="center")
            ws_log.cell(row=log_r, column=3, value=first_in).alignment = Alignment(horizontal="center")
            ws_log.cell(row=log_r, column=4, value=last_out).alignment = Alignment(horizontal="center")
            ws_log.cell(row=log_r, column=5, value=len(sessions)).alignment = Alignment(horizontal="center")
            ws_log.cell(row=log_r, column=6, value=break_mins).alignment = Alignment(horizontal="center")
            ws_log.cell(row=log_r, column=7, value=self._format_seconds_to_hm(work_sec)).alignment = Alignment(
                horizontal="center"
            )

            c_hrs = ws_log.cell(row=log_r, column=8, value=work_hrs)
            c_hrs.number_format = "0.00"

            c_tgt = ws_log.cell(row=log_r, column=9, value=day_target)
            c_tgt.number_format = "0.00"

            c_var = ws_log.cell(row=log_r, column=10, value=variance)
            c_var.number_format = "+0.00;-0.00;0.00"

            c_st = ws_log.cell(row=log_r, column=11, value=status_text)
            c_st.alignment = Alignment(horizontal="center")

            combined_notes = " | ".join(notes_list) if notes_list else ""
            ws_log.cell(row=log_r, column=12, value=combined_notes).alignment = Alignment(horizontal="left")

            # Base Row Styling
            base_row_fill = PatternFill(
                start_color=C_CARD_BG if is_weekend else (C_ZEBRA if log_r % 2 == 0 else "FFFFFF"),
                end_color=C_CARD_BG if is_weekend else (C_ZEBRA if log_r % 2 == 0 else "FFFFFF"),
                fill_type="solid",
            )
            for c in range(1, 13):
                cl = ws_log.cell(row=log_r, column=c)
                cl.font = font_regular
                cl.border = border_all
                cl.fill = base_row_fill

            # Status Badge Highlight
            if status_cat == "green":
                c_st.fill = PatternFill(start_color=C_GREEN_BG, end_color=C_GREEN_BG, fill_type="solid")
                c_st.font = Font(name="Segoe UI", size=10, bold=True, color=C_GREEN_FG)
            elif status_cat == "yellow":
                c_st.fill = PatternFill(start_color=C_YELLOW_BG, end_color=C_YELLOW_BG, fill_type="solid")
                c_st.font = Font(name="Segoe UI", size=10, bold=True, color=C_YELLOW_FG)
            elif status_cat == "red" and not is_weekend:
                c_st.fill = PatternFill(start_color=C_RED_BG, end_color=C_RED_BG, fill_type="solid")
                c_st.font = Font(name="Segoe UI", size=10, bold=True, color=C_RED_FG)

            log_r += 1

        # Totals / Summary Row
        ws_log.cell(row=log_r, column=1, value="TOTALS").font = font_bold
        ws_log.cell(row=log_r, column=1).alignment = Alignment(horizontal="center")
        ws_log.cell(row=log_r, column=5, value=f"=SUM(E4:E{log_r-1})")
        ws_log.cell(row=log_r, column=6, value=f"=SUM(F4:F{log_r-1})")
        ws_log.cell(row=log_r, column=8, value=f"=SUM(H4:H{log_r-1})").number_format = "0.00"
        ws_log.cell(row=log_r, column=9, value=f"=SUM(I4:I{log_r-1})").number_format = "0.00"
        ws_log.cell(row=log_r, column=10, value=f"=H{log_r}-I{log_r}").number_format = "+0.00;-0.00;0.00"

        for c in range(1, 13):
            cl = ws_log.cell(row=log_r, column=c)
            cl.font = font_bold
            cl.border = border_top_thick
            cl.fill = PatternFill(start_color=C_CARD_BG, end_color=C_CARD_BG, fill_type="solid")

        # Clean up default sheet if empty
        if default_sheet in wb.worksheets:
            wb.remove(default_sheet)

        # Auto-adjust column widths for both sheets
        for sheet in wb.worksheets:
            for col in sheet.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    # ignore merged title row length to avoid wide cols
                    if cell.row in (1, 2) and sheet.title in ("Monthly Summary", "Daily Time Logs"):
                        continue
                    val_str = str(cell.value or "")
                    if "\n" in val_str:
                        lines = val_str.split("\n")
                        max_len = max(max_len, max(len(l) for l in lines))
                    else:
                        max_len = max(max_len, len(val_str))
                sheet.column_dimensions[col_letter].width = max(max_len + 4, 12)

        wb.save(str(output_path))
        return output_path


# Global singleton instance
excel_exporter = ExcelReportExporter()
